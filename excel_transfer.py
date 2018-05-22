# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FILE_PATH = "./2016year.xlsx"
FROM_SHEET = "Sheet2"
TO_SHEET = "Sheet3"

STOCK_SIGN = u'证券代码'
WEEK_SIGN = u'交易周份'

book = None


def get_weeks(rows):
    we = []
    for row in rows:
        head_sign = row[0].value
        if head_sign and head_sign.replace("'", "") == STOCK_SIGN:
            for cell in row:
                try:
                    int(cell.value)
                    we.append(str(cell.value))
                except Exception, e:
                    pass
            return we


def get_week_count(row):
    for i in range(len(row)):
        cell = row[i]
        cv = cell.value
        if cv and cv == WEEK_SIGN:
            return i


def get_rows_value(weeks, rows):
    week_count = None
    result = {}

    for row in rows:
        try:
            if week_count is None:
                week_count = get_week_count(row)
            else:
                stock_code = row[week_count-1].value
                w = row[week_count].value
                val = row[week_count+1].value
                if result.get(stock_code, None) is None:
                    result[stock_code] = {}

                if str(w) not in weeks:
                    result[stock_code][str(w)] = ""
                else:
                    result[stock_code][str(w)] = val
        except Exception, e:
            pass
    return result


def write_values(values, to_sheet):
    keys = values.keys()

    def a(row):
        code = row[0].value
        if code and str(code) in keys:
            vals = values[code]
            for cell in row:
                if cell.value:
                    continue
                idx = str(cell.col_idx)
                if len(idx) < 2:
                    idx = "0" + idx
                v = vals.get(idx, None)
                cell.value = v

    for each in to_sheet.rows:
        try:
            a(each)
        except Exception, e:
            pass


def read():
    try:
        global book
        print "Start to read "
        book = load_workbook(FILE_PATH)

        from_sheet, to_sheet = book[FROM_SHEET], book[TO_SHEET]

        from_rows = from_sheet.rows
        columns = from_sheet.columns

        print "process weeks..."
        weeks = get_weeks(to_sheet.rows)

        print "process values..."
        result = get_rows_value(weeks, from_rows)

        print "write values..."
        write_values(result, to_sheet)

        print "read to save..."
        book.save("./output.xlsx")
        print "Success!"
    except Exception, e:
        import traceback
        print traceback.format_exc()

read()


wb = Workbook()

ws = wb.active

ws.cell(row=1, column=1).value = 6
ws.cell("B1").value = 7

for row in range(2, 11):
    for col in range(1, 11):
        ws.cell(row=row, column=col).value = get_column_letter(col)

ws.append(["我", "你", "她"])

wb.save(filename="/Users/budong/Desktop/a.xlsx")
